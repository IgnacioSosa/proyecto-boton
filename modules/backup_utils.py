import pandas as pd
import io
import json
import ast
import streamlit as st
from sqlalchemy import text
from .database import get_connection, get_engine, log_sql_error, ensure_clientes_schema, ensure_projects_schema, ensure_cliente_solicitudes_schema
from .utils import format_registro_date_iso, format_registro_datetime_iso


def _normalize_json_value(value):
    """Normaliza un valor JSON para garantizar que sea un string JSON válido con comillas dobles."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    
    # Si ya es un string, intentar parsearlo
    if isinstance(value, str):
        value_stripped = value.strip()
        if not value_stripped or value_stripped.lower() in ("nan", "nat", "none", "null"):
            return None
        
        try:
            # Intentar parsear como JSON válido primero
            parsed = json.loads(value_stripped)
            return json.dumps(parsed)
        except (json.JSONDecodeError, TypeError):
            try:
                # Intentar parsear como dict de Python (comillas simples)
                parsed = ast.literal_eval(value_stripped)
                return json.dumps(parsed)
            except (ValueError, SyntaxError, TypeError):
                # Si no se puede parsear, devolver el valor original (no es JSON)
                return value
    
    # Si es un diccionario o lista, serializar a JSON
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value)
    
    return value

pd.set_option('future.no_silent_downcasting', True)

SPANISH_MONTH_TO_NUMBER = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def _normalize_month_value(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    raw = str(value).strip()
    if not raw:
        return None
    raw_lower = raw.lower()
    if raw_lower in SPANISH_MONTH_TO_NUMBER:
        return SPANISH_MONTH_TO_NUMBER[raw_lower]
    digits = "".join(ch for ch in raw_lower if ch.isdigit())
    if digits:
        try:
            month_i = int(digits)
            if 1 <= month_i <= 12:
                return month_i
        except Exception:
            pass
    return value


def _is_date_only_column(column_name):
    name = str(column_name or "").strip().lower()
    if not name:
        return False
    if name == "fecha":
        return True
    if name.startswith("fecha_"):
        return True
    if name in {"fecha cierre", "fecha_cierre", "fecha ref", "fecha_ref", "fecha prevista", "fecha_prevista"}:
        return True
    return False


def _is_datetime_column(column_name):
    name = str(column_name or "").strip().lower()
    if not name:
        return False
    if name.endswith("_at"):
        return True
    if any(token in name for token in ("created_at", "updated_at", "processed_at", "last_login", "timestamp")):
        return True
    return False


def _parse_datetime_value(value, dayfirst=True):
    if value is None:
        return pd.NaT
    try:
        if pd.isna(value):
            return pd.NaT
    except Exception:
        pass

    # Si ya viene tipado como fecha/datetime, dejar que pandas lo normalice.
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        try:
            return pd.to_datetime(value, errors="coerce")
        except Exception:
            return pd.NaT

    raw = str(value).strip()
    if not raw:
        return pd.NaT

    # Priorizar formatos explícitos para no invertir mes/día en strings ISO.
    explicit_formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%y %H:%M:%S",
    ]
    for fmt in explicit_formats:
        try:
            return pd.to_datetime(raw, format=fmt, errors="raise")
        except Exception:
            pass

    try:
        return pd.to_datetime(raw, dayfirst=dayfirst, errors="coerce", utc=False)
    except Exception:
        return pd.NaT


def _coerce_datetime_series(series, dayfirst=True):
    parsed = series.apply(lambda v: _parse_datetime_value(v, dayfirst=dayfirst))
    try:
        if pd.api.types.is_datetime64tz_dtype(parsed):
            parsed = parsed.dt.tz_convert(None)
    except Exception:
        pass
    return parsed


def _normalize_registros_fecha_for_restore(df):
    if df is None or df.empty or 'fecha' not in df.columns:
        return df

    df = df.copy()

    def _resolve_fecha(row):
        raw = row.get('fecha')
        if raw is None:
            return None
        try:
            if pd.isna(raw):
                return None
        except Exception:
            pass

        raw_str = str(raw).strip()
        if not raw_str:
            return None

        # ISO/timestamp ISO: no ambiguo.
        parsed_iso = _parse_datetime_value(raw_str, dayfirst=True)
        if '-' in raw_str and pd.notna(parsed_iso):
            return parsed_iso.date().isoformat()

        # Fecha con "/" puede venir ambigua desde backups viejos/prod.
        if '/' in raw_str:
            parsed_dm = pd.NaT
            parsed_md = pd.NaT
            for fmt in ("%d/%m/%Y", "%d/%m/%y"):
                try:
                    parsed_dm = pd.to_datetime(raw_str, format=fmt, errors="raise")
                    break
                except Exception:
                    pass
            for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                try:
                    parsed_md = pd.to_datetime(raw_str, format=fmt, errors="raise")
                    break
                except Exception:
                    pass

            mes_hint = _normalize_month_value(row.get('mes'))
            try:
                mes_hint = int(mes_hint) if mes_hint is not None else None
            except Exception:
                mes_hint = None

            created_hint = _parse_datetime_value(row.get('created_at'), dayfirst=True)
            created_date = created_hint.date() if pd.notna(created_hint) else None

            candidates = []
            if pd.notna(parsed_dm):
                score = 0
                if mes_hint == int(parsed_dm.month):
                    score += 3
                if created_date and parsed_dm.date() == created_date:
                    score += 2
                candidates.append((score, 1, parsed_dm))
            if pd.notna(parsed_md):
                score = 0
                if mes_hint == int(parsed_md.month):
                    score += 3
                if created_date and parsed_md.date() == created_date:
                    score += 2
                candidates.append((score, 0, parsed_md))

            if candidates:
                candidates.sort(reverse=True)
                return candidates[0][2].date().isoformat()

        return parsed_iso.date().isoformat() if pd.notna(parsed_iso) else None

    df['fecha'] = df.apply(_resolve_fecha, axis=1)
    return df


def _normalize_dataframe_for_backup(df, column_types=None):
    if df is None or df.empty:
        return df

    df = df.copy()

    for col in df.columns:
        col_lower = str(col or "").strip().lower()
        if col_lower == "mes":
            df[col] = df[col].apply(_normalize_month_value)
            continue

        if _is_date_only_column(col):
            parsed = _coerce_datetime_series(df[col], dayfirst=True)
            df[col] = parsed.apply(lambda v: format_registro_date_iso(v, empty_value=None))
            continue

        if _is_datetime_column(col):
            parsed = _coerce_datetime_series(df[col], dayfirst=True)
            df[col] = parsed.apply(lambda v: format_registro_datetime_iso(v, empty_value=None))
            continue

        # Normalizar columnas JSON/JSONB (si tenemos información de tipos)
        if column_types and col in column_types and column_types[col] in ('json', 'jsonb'):
            df[col] = df[col].apply(_normalize_json_value)
            continue

        # También intentar normalizar cualquier columna que parezca contener dicts/lista
        # como fallback, para capturar casos donde no tenemos info de tipos
        try:
            # Verificar si hay algún valor que sea dict/list o string que parezca dict/list
            sample = df[col].dropna().iloc[0] if not df[col].dropna().empty else None
            if sample is not None:
                if isinstance(sample, (dict, list, tuple)):
                    df[col] = df[col].apply(_normalize_json_value)
                elif isinstance(sample, str):
                    sample_stripped = sample.strip()
                    if (sample_stripped.startswith('{') and sample_stripped.endswith('}')) or (sample_stripped.startswith('[') and sample_stripped.endswith(']')):
                        df[col] = df[col].apply(_normalize_json_value)
        except Exception:
            pass

    return df


def _apply_excel_formats(worksheet, df):
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    if df is None or df.empty:
        return

    date_columns = []
    datetime_columns = []

    for idx, col in enumerate(df.columns, start=1):
        if _is_date_only_column(col):
            date_columns.append(idx)
        elif _is_datetime_column(col):
            datetime_columns.append(idx)

    max_row = worksheet.max_row
    for col_idx in date_columns:
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, max_row + 1):
            worksheet[f"{col_letter}{row_idx}"].number_format = "DD/MM/YYYY"

    for col_idx in datetime_columns:
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, max_row + 1):
            worksheet[f"{col_letter}{row_idx}"].number_format = "DD/MM/YYYY HH:MM:SS"


def create_full_backup_excel():
    """Genera un archivo Excel con todas las tablas de la base de datos"""
    conn = get_connection()
    output = io.BytesIO()
    
    try:
        # Obtener lista de tablas públicas
        cursor = conn.cursor()
        cursor.execute("""
            SELECT tablename 
            FROM pg_catalog.pg_tables 
            WHERE schemaname = 'public'
        """)
        tables = [row[0] for row in cursor.fetchall()]
        tables.sort() # Orden alfabético para consistencia visual
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for table in tables:
                try:
                    cursor.execute(
                        """
                        SELECT column_name, data_type
                        FROM information_schema.columns
                        WHERE table_name = %s
                        """,
                        (table,),
                    )
                    col_info = cursor.fetchall()
                    cols = {row[0] for row in col_info}
                    column_types = {row[0]: row[1] for row in col_info}
                    order_clause = ""
                    if "id" in cols:
                        order_clause = ' ORDER BY "id"'
                    elif "created_at" in cols:
                        order_clause = ' ORDER BY "created_at"'

                    # Leer tabla (con orden estable cuando es posible)
                    engine = get_engine()
                    df = pd.read_sql_query(text(f'SELECT * FROM "{table}"{order_clause}'), con=engine)
                    df = _normalize_dataframe_for_backup(df, column_types)
                    
                    # Nombre de hoja (max 31 chars)
                    sheet_name = table[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    try:
                        ws = writer.sheets.get(sheet_name)
                        if ws is not None:
                            _apply_excel_formats(ws, df)
                    except Exception:
                        pass
                except Exception as e:
                    log_sql_error(f"Error exportando tabla {table}: {e}")
                    
        output.seek(0)
        return output
    except Exception as e:
        log_sql_error(f"Error generando backup: {e}")
        return None
    finally:
        conn.close()

def restore_full_backup_excel(uploaded_file, progress_callback=None):
    """Restaura la base de datos desde un archivo Excel.

    Args:
        uploaded_file: BytesIO o path-like con el Excel del backup.
        progress_callback (callable, optional): Función `(step_label, pct_0_1)`
            para reportar progreso a la UI. Opcional.

    Returns:
        (success: bool, msg: str)
    """
    def _report(label, pct):
        if progress_callback is None:
            return
        try:
            progress_callback(str(label), float(pct))
        except Exception:
            pass

    # Asegurar que el esquema esté actualizado antes de restaurar (columnas nuevas, tablas, etc.)
    try:
        ensure_clientes_schema()
        ensure_projects_schema()
        ensure_cliente_solicitudes_schema()
    except Exception as e:
        log_sql_error(f"Warning updating schema before restore: {e}")

    _report("Actualizando esquema...", 0.02)

    conn = get_connection()
    conn.autocommit = False # Usar transacción explícita
    cursor = conn.cursor()

    def _list_table_fks(conn_ref, table_name):
        cur_local = conn_ref.cursor()
        try:
            cur_local.execute(
                """
                SELECT tc.constraint_name,
                       kcu.table_name        AS child_table,
                       kcu.column_name       AS child_column,
                       ccu.table_name       AS parent_table,
                       ccu.column_name      AS parent_column
                FROM information_schema.table_constraints tc
                JOIN information_schema.key_column_usage kcu
                  ON tc.constraint_name = kcu.constraint_name
                 AND tc.table_schema   = kcu.table_schema
                JOIN information_schema.constraint_column_usage ccu
                  ON ccu.constraint_name = tc.constraint_name
                 AND ccu.table_schema    = tc.table_schema
                WHERE tc.constraint_type = 'FOREIGN KEY'
                  AND tc.table_schema    = 'public'
                  AND (kcu.table_name = %s OR ccu.table_name = %s)
                """,
                (str(table_name), str(table_name)),
            )
            return cur_local.fetchall() or []
        finally:
            try:
                cur_local.close()
            except Exception:
                pass

    def _drop_all_fks_and_snapshot(conn_ref, target_tables):
        cur_local = conn_ref.cursor()
        target_set = {str(t) for t in (target_tables or [])}
        fk_ddls = []
        try:
            cur_local.execute(
                """
                SELECT tc.table_name, tc.constraint_name
                FROM information_schema.table_constraints tc
                WHERE tc.constraint_type = 'FOREIGN KEY'
                  AND tc.table_schema    = 'public'
                """
            )
            rows = cur_local.fetchall() or []
            total = max(1, len(rows))
            for idx, (child_table, constraint_name) in enumerate(rows):
                if child_table not in target_set:
                    continue
                try:
                    cur_local.execute(
                        "SELECT pg_get_constraintdef(oid) "
                        "FROM pg_constraint "
                        "WHERE conname = %s AND conrelid = %s::regclass AND contype = 'f'",
                        (constraint_name, child_table),
                    )
                    res = cur_local.fetchone()
                    def_ddl = res[0] if res and res[0] else None
                    if def_ddl:
                        full_ddl = (
                            f'ALTER TABLE "{child_table}" '
                            f'ADD CONSTRAINT "{constraint_name}" {def_ddl}'
                        )
                        fk_ddls.append((child_table, constraint_name, full_ddl))
                except Exception:
                    pass
                try:
                    cur_local.execute(f'ALTER TABLE "{child_table}" DROP CONSTRAINT IF EXISTS "{constraint_name}"')
                except Exception:
                    pass
                if (idx + 1) % max(1, total // 10) == 0 or (idx + 1) == total:
                    conn_ref.commit()
            return fk_ddls
        finally:
            try:
                cur_local.close()
            except Exception:
                pass

    def _restore_fks_from_snapshot(conn_ref, fk_ddls_snapshot):
        cur_local = conn_ref.cursor()
        try:
            for child_table, constraint_name, full_ddl in (fk_ddls_snapshot or []):
                if not full_ddl:
                    continue
                try:
                    cur_local.execute(
                        f'ALTER TABLE "{child_table}" DROP CONSTRAINT IF EXISTS "{constraint_name}"'
                    )
                except Exception:
                    pass
                try:
                    cur_local.execute(full_ddl)
                except Exception as e:
                    log_sql_error(f"Warning re-creating FK {child_table}.{constraint_name}: {e}")
            conn_ref.commit()
        finally:
            try:
                cur_local.close()
            except Exception:
                pass
    
    # Orden de eliminación (Tablas hijas primero para evitar FK constraint errors)
    # IMPORTANTE: Mantener este orden sincronizado con las relaciones de la BD
    DELETE_ORDER = [
        'proyecto_documentos',
        'proyecto_compartidos',
        'activity_logs',
        'registros',      # Depende de: tecnicos, clientes, tipos_tarea, modalidades
        'nomina',         # Depende de: roles (departamento)
        'proyectos',      # Depende de: clientes, contactos (moved up to fix FK issue)
        'tecnicos',       # Tabla base para FK de registros (registros_id_tecnico_fkey)
        'contactos',      # Puede depender de clientes
        'clientes',       # Base para registros, proyectos, contactos
        'tipos_tarea',    # Base para registros (registros_id_tipo_fkey)
        'modalidades_tarea', # Base para registros (registros_id_modalidad_fkey) - Nombre correcto en BD
        'marcas',
        'usuarios',
        'roles',
        'grupos',
        'licencias'
    ]

    engine = get_engine()

    def _clean_df_for_table(table_name, df_raw):
        """Limpia y normaliza un DataFrame para insertar en una tabla.

        Devuelve el DataFrame limpio o None si no hay columnas válidas / está vacío.
        """
        if df_raw is None or len(df_raw) == 0:
            return None

        df_clean = df_raw.astype(object)

        if table_name == 'registros':
            df_clean = _normalize_registros_fecha_for_restore(df_clean)

        df_clean.replace(["NaT", "nan", "NaN"], None, inplace=True)
        df_clean = df_clean.where(pd.notnull(df_clean), None)

        try:
            cursor.execute(f"""
                SELECT column_name, is_nullable, data_type
                FROM information_schema.columns
                WHERE table_name = '{table_name}'
            """)
            schema = {row[0]: {'nullable': row[1] == 'YES', 'type': row[2]} for row in cursor.fetchall()}

            allowed_columns = [c for c in df_clean.columns if c in schema]
            if not allowed_columns:
                return None
            if len(allowed_columns) != len(df_clean.columns):
                df_clean = df_clean[allowed_columns]

            for col in df_clean.columns:
                if col in schema:
                    props = schema[col]
                    if props['type'] in ('json', 'jsonb'):
                        df_clean[col] = df_clean[col].apply(_normalize_json_value)
                    elif not props['nullable']:
                        if props['type'] in ('character varying', 'text', 'character', 'bpchar'):
                            df_clean[col] = df_clean[col].fillna('')
                        elif props['type'] in ('integer', 'bigint', 'smallint', 'numeric', 'double precision', 'real'):
                            df_clean[col] = df_clean[col].fillna(0)
                        elif props['type'] == 'boolean':
                            df_clean[col] = df_clean[col].fillna(False)

            if table_name == 'registros' and 'fecha' in df_clean.columns:
                from datetime import date as _date

                def _is_missing_fecha(v):
                    if v is None:
                        return True
                    if isinstance(v, str) and not v.strip():
                        return True
                    try:
                        return bool(pd.isna(v))
                    except Exception:
                        return False

                df_clean['fecha'] = df_clean['fecha'].apply(
                    lambda v: (
                        _parse_datetime_value(v, dayfirst=True).date().isoformat()
                        if pd.notna(_parse_datetime_value(v, dayfirst=True))
                        else None
                    )
                )

                missing_mask = df_clean['fecha'].apply(_is_missing_fecha)
                if bool(missing_mask.any()):
                    if 'created_at' in df_clean.columns:
                        created_parsed = pd.to_datetime(df_clean['created_at'], errors='coerce')
                        created_dates = created_parsed.dt.date
                        df_clean.loc[missing_mask, 'fecha'] = created_dates.loc[missing_mask].apply(
                            lambda d: d.isoformat() if d else None
                        )
                    still_missing = df_clean['fecha'].apply(_is_missing_fecha)
                    if bool(still_missing.any()):
                        df_clean.loc[still_missing, 'fecha'] = _date.today().isoformat()

            df_clean = df_clean.where(pd.notnull(df_clean), None)
        except Exception as e:
            log_sql_error(f"Warning checking schema for {table_name}: {e}")

        return df_clean

    def _bulk_insert_table(table_name, df_clean):
        """Inserta un DataFrame limpio usando to_sql bulk (método multi) con fallback executemany."""
        if df_clean is None or len(df_clean) == 0:
            return

        rows = len(df_clean)
        try:
            df_clean.to_sql(
                name=table_name,
                con=engine,
                if_exists='append',
                index=False,
                method='multi',
                chunksize=2000,
            )
            return
        except Exception as bulk_err:
            log_sql_error(
                f"Warning bulk insert {table_name} ({rows} rows) fallback a executemany: {bulk_err}"
            )

        columns = list(df_clean.columns)
        cols_str = ",".join([f'"{c}"' for c in columns])
        placeholders = ",".join(["%s"] * len(columns))
        query = f'INSERT INTO "{table_name}" ({cols_str}) VALUES ({placeholders})'
        values = df_clean.values.tolist()

        chunk = 500
        for i in range(0, len(values), chunk):
            batch = values[i:i + chunk]
            cursor.executemany(query, batch)
            conn.commit()

    try:
        _report("Leyendo Excel...", 0.04)
        xls = pd.read_excel(uploaded_file, sheet_name=None, na_values=['NaT'])

        cursor.execute("SELECT tablename FROM pg_catalog.pg_tables WHERE schemaname = 'public'")
        db_tables = [row[0] for row in cursor.fetchall()]

        _report("Quitando constraints...", 0.08)
        FK_RESTORE_SNAPSHOT = []
        try:
            FK_RESTORE_SNAPSHOT = _drop_all_fks_and_snapshot(conn, db_tables)
        except Exception as e:
            log_sql_error(f"Warning drop FKs pre-restore: {e}")

        _report("Borrando datos...", 0.14)
        processed_deletes = set()

        for table in DELETE_ORDER:
            if table in db_tables:
                cursor.execute(f"TRUNCATE TABLE {table} CASCADE")
                processed_deletes.add(table)

        for table in db_tables:
            if table not in processed_deletes:
                cursor.execute(f"TRUNCATE TABLE {table} CASCADE")
        conn.commit()

        INSERT_ORDER = list(reversed(DELETE_ORDER))

        sheet_targets = []
        processed_inserts = set()

        for table in INSERT_ORDER:
            sheet_name = table[:31]
            if sheet_name in xls:
                sheet_targets.append((sheet_name, table))
                processed_inserts.add(sheet_name)

        for sheet_name, df in xls.items():
            if sheet_name not in processed_inserts:
                target_table = None
                if sheet_name in db_tables:
                    target_table = sheet_name
                else:
                    for t in db_tables:
                        if t[:31] == sheet_name:
                            target_table = t
                            break
                if target_table:
                    sheet_targets.append((sheet_name, target_table))

        total_sheets = max(1, len(sheet_targets))
        insert_start_pct = 0.20
        insert_end_pct = 0.78
        pct_per_sheet = (insert_end_pct - insert_start_pct) / total_sheets

        for idx, (sheet_name, target_table) in enumerate(sheet_targets):
            current_pct = insert_start_pct + idx * pct_per_sheet
            _report(f"Insertando {target_table}...", current_pct)
            df_raw = xls.get(sheet_name)
            cleaned = _clean_df_for_table(target_table, df_raw)
            if cleaned is not None and len(cleaned) > 0:
                _bulk_insert_table(target_table, cleaned)
        conn.commit()

        _report("Ajustando secuencias...", 0.82)
        for table in db_tables:
            cursor.execute(f"""
                SELECT column_name, column_default
                FROM information_schema.columns
                WHERE table_name = '{table}'
                AND column_default LIKE 'nextval%%'
            """)
            serial_cols = cursor.fetchall()

            for col_name, col_default in serial_cols:
                try:
                    cursor.execute(f"SELECT pg_get_serial_sequence('{table}', '{col_name}')")
                    seq_res = cursor.fetchone()
                    if seq_res and seq_res[0]:
                        seq_name = seq_res[0]
                        cursor.execute(f"""
                            SELECT setval('{seq_name}', (SELECT COALESCE(MAX("{col_name}"), 0) + 1 FROM "{table}"), false)
                        """)
                except Exception as e:
                    log_sql_error(f"Warning reset sequence {table}.{col_name}: {e}")
        conn.commit()

        _report("Restaurando FKs...", 0.90)
        try:
            _restore_fks_from_snapshot(conn, FK_RESTORE_SNAPSHOT)
        except Exception as e:
            log_sql_error(f"Warning re-create FKs post-restore: {e}")

        _report("Saneos finales...", 0.96)
        try:
            from .database import (
                bootstrap_missing_roles_and_users,
                migrate_task_type_department_roles,
                repair_task_type_roles_missing_from_departments,
                repair_task_types_without_any_roles,
            )
            # Paso 0: asegurar que existan tecnico/comercial/compras y que los
            # usuarios no-admin tengan rol individual (no dpto_* ni adm_*).
            bootstrap_missing_roles_and_users()
            migrate_task_type_department_roles()
            repair_task_type_roles_missing_from_departments()
            repair_task_types_without_any_roles()
        except Exception as e:
            log_sql_error(f"Warning post-restore saneos tipos_tarea_roles: {e}")

        conn.commit()
        _report("Finalizado.", 1.0)
        return True, "Restauración completada exitosamente. Todas las tablas han sido recargadas."

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f"Error crítico en restauración: {str(e)}"
    finally:
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
